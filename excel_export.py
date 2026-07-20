"""
==========================================================
Flowering Synchronisation Analysis Tool
excel_export.py - Excel writer with 3 worksheets

Sheets:
  1. All Comparisons  - every isolation/surrounding pair analyzed
  2. Flowering Sync Only - Remarks == "Crop Flowering Sync"
  3. Summary - key statistics
==========================================================
"""

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _format_sheet(ws, df):
    """Freeze header row, add filters, auto-adjust column widths, style header."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def export_excel(results, stats, output_path):
    """
    results : list of dict rows (from analysis engine)
    stats   : dict of summary statistics
    output_path : path to .xlsx file
    """
    all_df = pd.DataFrame(results)

    if all_df.empty:
        all_df = pd.DataFrame(columns=[
            "Isolation_ID", "Isolation_Crop", "Isolation_Start", "Isolation_End",
            "Surrounding_ID", "Surrounding_Crop", "Surrounding_Start", "Surrounding_End",
            "Distance_m", "Crop_Match", "Flower_Overlap",
            "Overlap_Start", "Overlap_End", "Overlap_Days", "Remarks"
        ])

    sync_df = all_df[all_df["Remarks"] == "Crop Flowering Sync"].copy()

    summary_rows = [{"Metric": k, "Value": v} for k, v in stats.items()]
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_df.to_excel(writer, sheet_name="All Comparisons", index=False)
        sync_df.to_excel(writer, sheet_name="Flowering Sync Only", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        _format_sheet(writer.sheets["All Comparisons"], all_df)
        _format_sheet(writer.sheets["Flowering Sync Only"], sync_df)
        _format_sheet(writer.sheets["Summary"], summary_df)

    return output_path
